import os.path
import transients
import numpy as np
import matplotlib.pylab as plt
from cellpose import models
from openpyxl import Workbook
from tkinter import filedialog
from matplotlib import rcParams

"""
Authors: Jacob Soto & Xuan Fang
Program requires about 8gb of memory
Usage: python -u "~C:\ filePath\ of\ this\ program.py~"
"""

class container():
  def __init__(self, fileName,
                    index = None,
                    raw = None,   # array of data
                    mask = None,  # array of cells locations
                    binaryMask = None  # binary array of cells within the img
                    ):
    self.fileName = fileName
    self.index = index
    self.raw = raw
    self.mask = mask
    self.binaryMask = binaryMask

class cellProp():
    def __init__(self, coords=None, area=None):
        self.coords = coords
        self.area = area

def exportData(sheet, dataSet, title, count):
  for i, sublist in enumerate(dataSet):
      for j, item in enumerate(sublist):
          header = sheet.cell(row = 1, column = (j*4)+count)
          header.value = (str(title) + str(j+1))

          cell = sheet.cell(row = i+2, column = (j*4)+count)
          cell.value = item

def main():
  path = filedialog.askdirectory()
  path += "/"
  fileNames = ["CFP.tif", "GFP.tif", "mCherry.tif"]
  channelNames = ["CFP", "GFP", "MCHERRY"]

  # create a dictionary to store channel info
  channels = {}
  for i, name in enumerate(channelNames):
      if (os.path.exists(path + fileNames[i])):
        channel = container(fileNames[i], index=i)
        channels[name] = channel
        channels[name].channel_index = i
      else:
        print("Could not find file: " + fileNames[i])
        exit(1)

  # references
  CFP = channels['CFP']
  GFP = channels['GFP']
  MCHERRY = channels['MCHERRY']


  # get raw data from images
  for channel in channels.values():
    print("processing " + channel.fileName)
    if (channel == MCHERRY):
      ar = transients.LoadTimeData(path + channel.fileName, timeReversed = True)
    else:
      ar = transients.LoadStaticData(path + channel.fileName, timeReversed = True)
    channel.raw = np.asarray(ar)
  print()

  # define cellpose model and run cellpose to find channel masks
  model = models.Cellpose(gpu=True, model_type='cyto2')

  for channel in channels.values():
      if (channel == MCHERRY):
          data = channel.raw[0, :, :]
      else:
          data = channel.raw[:, :]
      
      print("Processing mask for: " + channel.fileName)
      masks, flows, styles, diams = model.eval(data, diameter=None, do_3D=False, )
      channel.mask = masks


  # convert masks to binary format
  for key, channel in channels.items():
      channel.binaryMask = np.zeros_like(channel.mask, int)
      channel.binaryMask[channel.mask > 0] = 1


  # find cells only expressed in all three channels
  combo_BinaryMask = CFP.binaryMask + GFP.binaryMask + MCHERRY.binaryMask
  final_BinaryMask = np.zeros_like(combo_BinaryMask, int)
  final_BinaryMask[combo_BinaryMask == 3] = 1

  dataSet = final_BinaryMask * MCHERRY.raw
  CFP_values = final_BinaryMask * CFP.raw

  print("Processing mask for: data set")
  masks, flows, styles, diams = model.eval(dataSet[0, :, :], diameter=None, do_3D=False, )
  MCHERRY.mask = masks
  cellCount = np.max(final_BinaryMask * MCHERRY.mask)

  # cell indicies
  indices=[]
  areas=[]
  for i in range(1, cellCount+1):
      index = np.argwhere(MCHERRY.mask == i)
      area = len(index)
      indices.append(index)
      areas.append(area)
  
  indices = np.array(indices, dtype=object)
  areas = np.array(areas, dtype=object)

  region_cells = []
  for i in range(cellCount):
      prop = cellProp(indices[i], areas[i])
      region_cells.append(prop)

  MCHERRY.region_cells = region_cells


  # get time series data for selected cells in all three channels
  traces, region_cells_master = transients.GetTraces(
                dataSet,
                final_BinaryMask,
                region_cells = MCHERRY.region_cells,
                channelName = MCHERRY.index
                )
  dataSet = traces

  # get CFP values for selected cells
  traces, region_cells_master = transients.GetTraces(
                CFP_values,
                final_BinaryMask,
                region_cells = MCHERRY.region_cells,
                channelName = MCHERRY.index
                )
  CFP_values = traces

  # Normalize data
  normalized = [None] * len(dataSet)
  for i in range(len(dataSet)):
      trace = dataSet[i]
      minVal = np.min(trace)
      maxVal = np.max(trace)
      norm = (trace - minVal) / (maxVal - minVal)
      normalized[i] = norm

  # Transpose data for export
  dataSet = np.array(dataSet)
  normalized = np.array(normalized)
  transposed_Data = dataSet.T
  transposed_Norm = normalized.T

  # Save data
  wb = Workbook()
  sheet = wb.active
  exportData(sheet, transposed_Data, "Cell ", 1)
  exportData(sheet, transposed_Norm, "Norm ", 2)

  for j, item in enumerate(CFP_values):
      header = sheet.cell(row = 1, column = (j*4)+3)
      header.value = ("CFP " + str(j+1))
      
      cell = sheet.cell(row = 2, column = (j*4)+3)
      cell.value = item

  destFile = (path + "Cell_Data.xlsx")
  wb.save(destFile)

  # Save cell segmenation img
  plt.figure(figsize=(20,10))
  ax1 = plt.subplot(131)
  ax1.set_title('Selected Cells')
  ax1.imshow(MCHERRY.mask)
  labels=[]

  for i in range(MCHERRY.mask.shape[0]):
      for j in range(MCHERRY.mask.shape[1]):
          if (MCHERRY.mask[i][j] == 0):
              continue
          else:
              if MCHERRY.mask[i][j] not in labels:
                  labels.append(MCHERRY.mask[i][j])
                  ax1.text(j, i, int(MCHERRY.mask[i][j]), ha="center", va="center", fontsize=12, fontweight='black', color='orange')

  plt.savefig(path + 'Cell_Selection.png', bbox_inches='tight')
  print("Done, results save to: " + str(path))


main()