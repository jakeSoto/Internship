import time
import os.path
import transients
import numpy as np
import matplotlib.pylab as plt
from cellpose import models
from openpyxl import Workbook
from tkinter import filedialog
from multiprocessing import Pool, Process

class container():
  def __init__(self, fileName,
                    index = None,
                    raw = None,   # array of data
                    mask = None,  # array of cells locations
                    binaryMask = None  # binary array of cells within the img
                    ):
    self.fileName = fileName
    self.index = index
    self.raw = raw
    self.mask = mask
    self.binaryMask = binaryMask

class cellProp():
    def __init__(self, coords=None, area=None):
        self.coords = coords
        self.area = area

# For use with multiprocessing pool
def runCellpose(data):
    model = models.Cellpose(gpu=True, model_type='cyto2')
    print("Processing mask")
    masks, flows, styles, diams = model.eval(data, diameter=None, do_3D=False, )
    return masks


# Get channel masks
# Uses multiprocessing to reduce overhead
def multiProcess(channelDict: dict) -> dict:
    dataSet = []

    for channel in channelDict.values():
        if (channel.fileName == "mCherry.tif"):
            dataSet.append(channel.raw[0])
        else:
            dataSet.append(channel.raw)

    with Pool() as pool:
        results = pool.map(runCellpose, dataSet)

    for channel in channelDict.values():
        channel.mask = results[channel.index]

    return channelDict


# Creates a dictionary for channel data
def createDict(fileNames: str, path: str) -> dict:
    channelDict = {}

    for i, name in enumerate(fileNames):
        print("Processing channel: " + name)
        if (name == "mCherry.tif"):
            ar = transients.LoadTimeData(path + name, timeReversed = True)
        else:
            ar = transients.LoadStaticData(path + name, timeReversed = True)

        channelDict[name] = container(fileName=name, index=i, raw=ar)
        channelDict[name].index = i
        print()

    return channelDict


def exportData(sheet, dataSet, title, count):
    for i, sublist in enumerate(dataSet):
        for j, item in enumerate(sublist):
            header = sheet.cell(row = 1, column = (j*4)+count)
            header.value = (str(title) + str(j+1))

            cell = sheet.cell(row = i+2, column = (j*4)+count)
            cell.value = item


def main():
    start_time = time.time()
    path = filedialog.askdirectory()
    path += "/"
    fileNames = ["CFP.tif", "YFP.tif", "mCherry.tif"]

    channels = {}
    channels = createDict(fileNames, path)
    channels = multiProcess(channels)

    CFP = channels['CFP.tif']
    GFP = channels['YFP.tif']
    MCHERRY = channels['mCherry.tif']

    # Convert masks to binary
    for key, channel in channels.items():
        channel.binaryMask = np.zeros_like(channel.mask, int)
        channel.binaryMask[channel.mask > 0] = 1

    # Select cells
    combo_BinaryMask = CFP.binaryMask + GFP.binaryMask + MCHERRY.binaryMask
    MCHERRY.mask = runCellpose(combo_BinaryMask)

    final_BinaryMask = np.zeros_like(MCHERRY.mask, int)
    final_BinaryMask[MCHERRY.mask > 0] = 1

    dataSet = final_BinaryMask * MCHERRY.raw
    CFP_values = final_BinaryMask * CFP.raw
    cellCount = np.max(final_BinaryMask * MCHERRY.mask)

    # Cell indicies
    indices=[]
    areas=[]
    for i in range(1, cellCount+1):
        index = np.argwhere(MCHERRY.mask == i)
        area = len(index)
        indices.append(index)
        areas.append(area)
    
    indices = np.array(indices, dtype=object)
    areas = np.array(areas, dtype=object)

    region_cells = []
    for i in range(cellCount):
        prop = cellProp(indices[i], areas[i])
        region_cells.append(prop)

    MCHERRY.region_cells = region_cells

    # Time series data
    traces, region_cells_master = transients.GetTraces(
                    dataSet,
                    final_BinaryMask,
                    region_cells = MCHERRY.region_cells
                    )
    cellTraces = traces

    # CFP values
    traces, region_cells_master = transients.GetTraces(
                    CFP_values,
                    final_BinaryMask,
                    region_cells = MCHERRY.region_cells
                    )
    CFP_values = traces

    # Normalize data
    normalized = [None] * len(cellTraces)
    for i in range(len(cellTraces)):
        trace = cellTraces[i]
        minVal = np.min(trace)
        maxVal = np.max(trace)
        norm = (trace - minVal) / (maxVal - minVal)
        normalized[i] = norm

    # Transpose data for export
    cellTraces = np.array(cellTraces)
    normalized = np.array(normalized)
    transposed_Cells = cellTraces.T
    transposed_Norm = normalized.T

    # Save data
    wb = Workbook()
    sheet = wb.active
    exportData(sheet, transposed_Cells, "Cell ", 1)
    exportData(sheet, transposed_Norm, "Norm ", 2)

    for j, item in enumerate(CFP_values):
        header = sheet.cell(row = 1, column = (j*4)+3)
        header.value = ("CFP " + str(j+1))
        
        cell = sheet.cell(row = 2, column = (j*4)+3)
        cell.value = item

    destFile = (path + "Cell_Data.xlsx")
    wb.save(destFile)

    # Save cell segmenation img
    plt.figure(figsize=(20,10))
    ax1 = plt.subplot(131)
    ax1.set_title('Selected Cells')
    ax1.imshow(MCHERRY.mask)
    labels=[]

    for i in range(MCHERRY.mask.shape[0]):
        for j in range(MCHERRY.mask.shape[1]):
            if (MCHERRY.mask[i][j] == 0):
                continue
            else:
                if MCHERRY.mask[i][j] not in labels:
                    labels.append(MCHERRY.mask[i][j])
                    ax1.text(j, i, int(MCHERRY.mask[i][j]), ha="center", va="center", fontsize=12, fontweight='black', color='orange')

    plt.savefig(path + 'Cell_Selection.png', bbox_inches='tight')

    end_time = time.time()
    total_time = (end_time - start_time) / 60
    print("Done, results save to: " + str(path))
    print("Execution time: " + str(total_time))



if __name__ == "__main__":
    main()

