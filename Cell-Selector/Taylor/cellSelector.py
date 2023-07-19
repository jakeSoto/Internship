import os, time
import helper
from openpyxl import Workbook


def main():
    start_time = time.time()
    folderDict, dest = helper.searchDirectory()
    measuredName = helper.measuredChannel(folderDict)
    print("Measuring the " + str(measuredName) + " channel")

    imgRoot = os.path.join(dest, "Cells_")
    dest = os.path.join(dest, "Cell_Data.xlsx")
    channels = {}
    wb = Workbook()

    for i, folder in enumerate(folderDict):
        if (folderDict[folder] == []):
            continue

        # Get container objects for channels in each folder
        else:
            channels = helper.createChannelDict(folderDict[folder])
            sheet = wb.create_sheet(folder, i)

        # 2 Channels
        if (len(channels) < 3):
            channels = helper.processTwoChannels(channels, measuredName)

        # 3 Channels
        else:
            channels = helper.processThreeChannels(channels, measuredName)

        MCHERRY = channels['mCherry']
        MEASURED = channels[measuredName]
        normalized = helper.normalizeData(MCHERRY.traces)
        dividedVals, normalizedVals = helper.newNormalized(MCHERRY.traces, MEASURED.traces)

        # Export data
        helper.exportData(sheet, MCHERRY.traces, "Cell ", 1)
        helper.exportData(sheet, normalized, "Norm ", 2)

        for i, value in enumerate(MEASURED.traces):
            header = sheet.cell(row = 1, column = (i*6)+3)
            header.value = (measuredName + " " + str(i+1))

            cell = sheet.cell(row = 2, column = (i*6)+3)
            cell.value = value

        helper.exportData(sheet, dividedVals, ("mCherry/" + measuredName + " "), 4)
        helper.exportData(sheet, normalizedVals, "New Norm ", 5)

        imgName = imgRoot + str(folder) + ".png"
        helper.saveCellImg(MCHERRY.mask, imgName)

    wb.save(dest)
    end_time = time.time()
    total_time = (end_time - start_time) / 60
    print("Done, results save to: " + str(dest))
    print("Execution time: " + str(round(total_time, 1)) + " minutes")


if __name__ == "__main__":
    main()
