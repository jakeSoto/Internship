import time
import threading
import os.path
import transients
import numpy as np
import matplotlib.pylab as plt
from cellpose import models
from openpyxl import Workbook
from tkinter import filedialog
from matplotlib import rcParams




def main():
    path = filedialog.askdirectory()
    path += "/"

    # create a dictionary to store channel info
    channels = {}
    channels['STATIC'] = container("CFP.tif", index = 0)
    channels['STATIC'].channel_index = 0
    channels['MCHERRY'] = container("mCherry.tif", index = 1)
    channels['MCHERRY'].channel_index = 1

    # references
    STATIC = channels['STATIC']
    MCHERRY = channels['MCHERRY']

    # get raw data from images
    print("Processing " + STATIC.fileName)
    ar1 = transients.LoadStaticData(path + STATIC.fileName, timeReversed = True)
    STATIC.raw = np.asarray(ar1)
    print()
    print("Processing " + MCHERRY.fileName)
    ar2 = transients.LoadTimeData(path + MCHERRY.fileName, timeReversed = True)
    MCHERRY.raw = np.asarray(ar2)
    print()

    STATIC.mask = runCellpose(STATIC.raw)
    MCHERRY.mask = runCellpose(MCHERRY.raw[0])

    # convert masks to binary format
    STATIC.binaryMask = np.zeros_like(STATIC.mask, int)
    STATIC.binaryMask[STATIC.mask > 0] = 1

    MCHERRY.binaryMask = np.zeros_like(MCHERRY.mask, int)
    MCHERRY.binaryMask[MCHERRY.mask > 0] = 1


    # find cells only expressed in all three channels
    combo_BinaryMask = STATIC.binaryMask + MCHERRY.binaryMask
    MCHERRY.mask = runCellpose(combo_BinaryMask)


    final_BinaryMask = np.zeros_like(MCHERRY.mask, int)
    final_BinaryMask[MCHERRY.mask > 0] = 1

    dataSet = final_BinaryMask * MCHERRY.raw
    STATIC_values = final_BinaryMask * STATIC.raw
    cellCount = np.max(final_BinaryMask * MCHERRY.mask)

    # cell indicies
    indices=[]
    areas=[]
    for i in range(1, cellCount+1):
        index = np.argwhere(MCHERRY.mask == i)
        area = len(index)
        indices.append(index)
        areas.append(area)
    
    indices = np.array(indices, dtype=object)
    areas = np.array(areas, dtype=object)

    region_cells = []
    for i in range(cellCount):
        prop = cellProp(indices[i], areas[i])
        region_cells.append(prop)

    MCHERRY.region_cells = region_cells


    # get time series data for selected cells in all three channels
    traces, region_cells_master = transients.GetTraces(
                    dataSet,
                    final_BinaryMask,
                    region_cells = MCHERRY.region_cells,
                    channelName = MCHERRY.index
                    )
    dataSet = traces

    # get STATIC values for selected cells
    traces, region_cells_master = transients.GetTraces(
                    STATIC_values,
                    final_BinaryMask,
                    region_cells = MCHERRY.region_cells,
                    channelName = MCHERRY.index
                    )
    STATIC_values = traces

    # Normalize data
    normalized = [None] * len(dataSet)
    for i in range(len(dataSet)):
        trace = dataSet[i]
        minVal = np.min(trace)
        maxVal = np.max(trace)
        norm = (trace - minVal) / (maxVal - minVal)
        normalized[i] = norm

    # Transpose data for export
    dataSet = np.array(dataSet)
    normalized = np.array(normalized)
    transposed_Data = dataSet.T
    transposed_Norm = normalized.T

    # Save data
    wb = Workbook()
    sheet = wb.active
    exportData(sheet, transposed_Data, "Cell ", 1)
    exportData(sheet, transposed_Norm, "Norm ", 2)

    for j, item in enumerate(STATIC_values):
        header = sheet.cell(row = 1, column = (j*4)+3)
        header.value = ("GFP " + str(j+1))
        
        cell = sheet.cell(row = 2, column = (j*4)+3)
        cell.value = item

    destFile = (path + "Cell_Data.xlsx")
    wb.save(destFile)

    # Save cell segmenation img
    plt.figure(figsize=(20,10))
    ax1 = plt.subplot(131)
    ax1.set_title('Selected Cells')
    ax1.imshow(MCHERRY.mask)
    labels=[]

    for i in range(MCHERRY.mask.shape[0]):
        for j in range(MCHERRY.mask.shape[1]):
            if (MCHERRY.mask[i][j] == 0):
                continue
            else:
                if MCHERRY.mask[i][j] not in labels:
                    labels.append(MCHERRY.mask[i][j])
                    ax1.text(j, i, int(MCHERRY.mask[i][j]), ha="center", va="center", fontsize=12, fontweight='black', color='orange')

    plt.savefig(path + 'Cell_Selection.png', bbox_inches='tight')
    print("Done, results save to: " + str(path))


if __name__ == '__main__':
    main()